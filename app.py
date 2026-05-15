from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file)
from functools import wraps
import pymysql
import bcrypt
import json
import numpy as np
import cv2
import base64
import math
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from deepface import DeepFace
import warnings
warnings.filterwarnings('ignore')
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import pandas as pd
from pyngrok import ngrok


app = Flask(__name__)
app.secret_key = 'face_attendance_2026_secure_key'

# ─── SECURITY: Flask Session Config ───
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE']  = 'Lax'


# ─── CONFIG ──────────────────────────────────────────────────────────────────
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',              # ← Change your MySQL password
    'db': 'face_attendance',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

COLLEGE_LAT    = 12.887780 # ← Change to your college latitude
COLLEGE_LON    = 80.106281 # ← Change to your college longitude
GEOFENCE_RADIUS = 200       # meters


# ─── DEEPFACE HELPER (SYNC - Fixed) ──────────────────────────────────────────
def extract_deepface_embedding(image_rgb):
    """Extract 512-dim embedding using Facenet512 model (synchronous)"""
    try:
        embedding = DeepFace.represent(
            image_rgb,
            model_name="Facenet512",
            enforce_detection=False,
            detector_backend="skip"
        )[0]["embedding"]
        return np.array(embedding)
    except Exception:
        return None


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def get_db():
    return pymysql.connect(**DB_CONFIG)


def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi    = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (math.sin(dphi / 2) ** 2 +
         math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'student_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


# ─── STUDENT ROUTES ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


# ---------- Registration ----------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name        = request.form['name'].strip()
        register_no = request.form['register_no'].strip().upper()
        department  = request.form['department'].strip()
        email       = request.form['email'].strip().lower()
        password    = request.form['password']

        if len(password) < 6:
            return render_template('register.html',
                                   error='Password must be at least 6 characters!')

        pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute(
                    "INSERT INTO students (name,register_no,department,email,password_hash) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (name, register_no, department, email, pw_hash)
                )
                db.commit()
                session['student_id']   = cur.lastrowid
                session['student_name'] = name
                session['register_no']  = register_no
                session['department']   = department
            return redirect(url_for('capture_face'))
        except pymysql.err.IntegrityError:
            return render_template('register.html',
                                   error='Register No or Email already exists!')
        finally:
            db.close()
    return render_template('register.html')


# ---------- Face Dataset Creation ----------
@app.route('/capture_face')
@login_required
def capture_face():
    return render_template('capture_face.html')


@app.route('/save_face', methods=['POST'])
@login_required
def save_face():                          # ← FIXED: removed async
    images_data = request.json.get('images', [])
    encodings = []

    for img_data in images_data:
        try:
            img_b64   = img_data.split(',')[1]
            img_bytes = base64.b64decode(img_b64)
            nparr     = np.frombuffer(img_bytes, np.uint8)
            img       = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            rgb       = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

            emb = extract_deepface_embedding(rgb)  # ← FIXED: removed await
            if emb is not None:
                encodings.append(emb.tolist())
        except Exception:
            continue

    if len(encodings) < 5:
        return jsonify({'success': False,
                        'message': f'Only {len(encodings)} faces detected. Need 5+. Ensure good lighting.'})

    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("UPDATE students SET face_encodings=%s WHERE id=%s",
                        (json.dumps(encodings), session['student_id']))
            db.commit()
        return jsonify({'success': True,
                        'message': f'✅ {len(encodings)} face embeddings registered! (Facenet512)'})
    finally:
        db.close()


# ---------- Login ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'student_id' in session:
        return redirect(url_for('mark_attendance'))

    if request.method == 'POST':
        register_no = request.form['register_no'].strip().upper()
        password    = request.form['password']
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute("SELECT * FROM students WHERE register_no=%s", (register_no,))
                student = cur.fetchone()

            if student and bcrypt.checkpw(password.encode(),
                                          student['password_hash'].encode()):
                session['student_id']   = student['id']
                session['student_name'] = student['name']
                session['register_no']  = student['register_no']
                session['department']   = student['department']
                if not student['face_encodings']:
                    return redirect(url_for('capture_face'))
                return redirect(url_for('mark_attendance'))
            else:
                return render_template('login.html',
                                       error='Invalid Register Number or Password!')
        finally:
            db.close()
    return render_template('login.html')


# ---------- Mark Attendance ----------
@app.route('/mark_attendance')
@login_required
def mark_attendance():
    today = date.today()
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT * FROM attendance WHERE student_id=%s AND date=%s",
                        (session['student_id'], today))
            existing = cur.fetchone()

            cur.execute("""
                SELECT date, time, status FROM attendance
                WHERE student_id=%s ORDER BY date DESC LIMIT 10
            """, (session['student_id'],))
            history = cur.fetchall()
    finally:
        db.close()

    return render_template('mark_attendance.html',
                           student_name=session['student_name'],
                           register_no=session['register_no'],
                           department=session.get('department', ''),
                           already_marked=(existing is not None),
                           history=history,
                           college_lat=COLLEGE_LAT,
                           college_lon=COLLEGE_LON,
                           radius=GEOFENCE_RADIUS,
                           today=today.strftime('%d-%m-%Y'))


@app.route('/verify_attendance', methods=['POST'])
@login_required
def verify_attendance():                  # ← FIXED: removed async
    data       = request.json
    latitude   = float(data.get('latitude'))
    longitude  = float(data.get('longitude'))
    image_data = data.get('image')

    # Step 1: Geofence
    dist = haversine(latitude, longitude, COLLEGE_LAT, COLLEGE_LON)
    if dist > GEOFENCE_RADIUS:
        return jsonify({'success': False,
                        'message': f'❌ You are outside college campus!\n'
                                   f'Distance: {int(dist)} meters from campus.'})

    today = date.today()
    db = get_db()
    try:
        with db.cursor() as cur:
            # Step 2: Duplicate check
            cur.execute("SELECT id FROM attendance WHERE student_id=%s AND date=%s",
                        (session['student_id'], today))
            if cur.fetchone():
                return jsonify({'success': False,
                                'message': '⚠️ Attendance already marked for today!'})

            # Step 3: Fetch stored encodings
            cur.execute("SELECT face_encodings,name,register_no,department "
                        "FROM students WHERE id=%s", (session['student_id'],))
            student = cur.fetchone()

        if not student['face_encodings']:
            return jsonify({'success': False,
                            'message': 'Face not registered! Please register first.'})

        stored_embeddings = [np.array(e) for e in json.loads(student['face_encodings'])]

        # Step 4: DeepFace Processing
        img_bytes = base64.b64decode(image_data.split(',')[1])
        nparr     = np.frombuffer(img_bytes, np.uint8)
        img       = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        rgb       = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        embedding = extract_deepface_embedding(rgb)  # ← FIXED: removed await
        if embedding is None:
            return jsonify({'success': False,
                            'message': '❌ No face detected! Ensure proper lighting and try again.'})

        # Cosine similarity
        similarities = [np.dot(embedding, stored) /
                       (np.linalg.norm(embedding) * np.linalg.norm(stored))
                       for stored in stored_embeddings]

        max_similarity = max(similarities)
        confidence = round(max_similarity * 100, 1)

        if max_similarity > 0.50:
            now = datetime.now()
            with db.cursor() as cur:
                cur.execute("""
                    INSERT INTO attendance
                        (student_id, date, time, status, latitude, longitude)
                    VALUES (%s,%s,%s,'Present',%s,%s)
                """, (session['student_id'], today,
                      now.strftime('%H:%M:%S'), latitude, longitude))
                db.commit()

            return jsonify({
                'success': True,
                'message': '✅ Attendance Marked Successfully!',
                'details': {
                    'name':        student['name'],
                    'register_no': student['register_no'],
                    'department':  student['department'],
                    'date':        today.strftime('%d-%m-%Y'),
                    'time':        now.strftime('%I:%M %p'),
                    'status':      'Present',
                    'confidence':  f'{confidence}% (Facenet512)'
                }
            })
        else:
            return jsonify({'success': False,
                            'message': f'❌ Face Not Recognized! Confidence: {confidence:.1f}%'})
    finally:
        db.close()


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


# ─── ADMIN ROUTES ─────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if 'admin_id' in session:
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        db = get_db()
        try:
            with db.cursor() as cur:
                cur.execute("SELECT * FROM admins WHERE username=%s", (username,))
                admin = cur.fetchone()
            if admin and bcrypt.checkpw(password.encode(),
                                        admin['password_hash'].encode()):
                session['admin_id']       = admin['id']
                session['admin_username'] = admin['username']
                return redirect(url_for('admin_dashboard'))
            else:
                return render_template('admin/login.html',
                                       error='Invalid admin credentials!')
        finally:
            db.close()
    return render_template('admin/login.html')


@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    today = date.today()
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(*) as count FROM students")
            total_students = cur.fetchone()['count']

            cur.execute("SELECT COUNT(*) as count FROM attendance WHERE date=%s", (today,))
            present_today = cur.fetchone()['count']

            cur.execute("""
                SELECT a.time, a.status, s.name, s.register_no, s.department
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.date=%s ORDER BY a.time DESC
            """, (today,))
            today_records = cur.fetchall()
    finally:
        db.close()

    return render_template('admin/dashboard.html',
                           total_students=total_students,
                           present_today=present_today,
                           absent_today=total_students - present_today,
                           today_records=today_records,
                           today=today.strftime('%d-%m-%Y'))


@app.route('/admin/students')
@admin_required
def admin_students():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("""
                SELECT s.id, s.name, s.register_no, s.department, s.email,
                       s.face_encodings,
                       DATE_FORMAT(s.created_at,'%%d-%%m-%%Y') as joined,
                       COUNT(a.id) as total_present
                FROM students s
                LEFT JOIN attendance a ON s.id = a.student_id
                GROUP BY s.id
                ORDER BY s.register_no
            """)
            students = cur.fetchall()
    finally:
        db.close()
    return render_template('admin/students.html', students=students)


@app.route('/admin/reports', methods=['GET', 'POST'])
@admin_required
def admin_reports():
    filter_date = request.form.get('filter_date', date.today().strftime('%Y-%m-%d'))
    filter_dept = request.form.get('filter_dept', '')
    report_type = request.form.get('report_type', 'daily')
    dept_clause = "AND s.department = %s" if filter_dept else ""

    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT DISTINCT department FROM students ORDER BY department")
            departments = [r['department'] for r in cur.fetchall()]

            if report_type == 'monthly':
                month = filter_date[:7]
                query = f"""
                    SELECT s.name, s.register_no, s.department,
                           COUNT(a.id) as days_present
                    FROM students s
                    LEFT JOIN attendance a
                        ON s.id = a.student_id
                        AND DATE_FORMAT(a.date,'%%Y-%%m') = %s
                    WHERE 1=1 {dept_clause}
                    GROUP BY s.id, s.name, s.register_no, s.department
                    ORDER BY s.register_no
                """
                params = (month, filter_dept) if filter_dept else (month,)
            else:
                query = f"""
                    SELECT s.name, s.register_no, s.department,
                           DATE_FORMAT(a.date,'%%d-%%m-%%Y')  AS att_date,
                           TIME_FORMAT(a.time,'%%h:%%i %%p')  AS att_time,
                           COALESCE(a.status,'Absent')        AS status
                    FROM students s
                    LEFT JOIN attendance a
                        ON s.id = a.student_id AND a.date = %s
                    WHERE 1=1 {dept_clause}
                    ORDER BY s.register_no
                """
                params = (filter_date, filter_dept) if filter_dept else (filter_date,)

            cur.execute(query, params)
            records = cur.fetchall()
    finally:
        db.close()

    return render_template('admin/reports.html',
                           records=records,
                           departments=departments,
                           filter_date=filter_date,
                           filter_dept=filter_dept,
                           report_type=report_type)


@app.route('/admin/export_excel', methods=['POST'])
@admin_required
def export_excel():
    filter_date = request.form.get('filter_date', date.today().strftime('%Y-%m-%d'))
    filter_dept = request.form.get('filter_dept', '')
    report_type = request.form.get('report_type', 'daily')
    dept_clause = "AND s.department = %s" if filter_dept else ""

    db = get_db()
    try:
        with db.cursor() as cur:
            if report_type == 'monthly':
                month = filter_date[:7]
                query = f"""
                    SELECT s.name, s.register_no, s.department, COUNT(a.id) as days_present
                    FROM students s
                    LEFT JOIN attendance a
                        ON s.id=a.student_id AND DATE_FORMAT(a.date,'%%Y-%%m')=%s
                    WHERE 1=1 {dept_clause}
                    GROUP BY s.id ORDER BY s.register_no
                """
                params = (month, filter_dept) if filter_dept else (month,)
            else:
                query = f"""
                    SELECT s.name, s.register_no, s.department,
                           COALESCE(DATE_FORMAT(a.date,'%%d-%%m-%%Y'),'') AS att_date,
                           COALESCE(TIME_FORMAT(a.time,'%%h:%%i %%p'),'') AS att_time,
                           COALESCE(a.status,'Absent') AS status
                    FROM students s
                    LEFT JOIN attendance a ON s.id=a.student_id AND a.date=%s
                    WHERE 1=1 {dept_clause}
                    ORDER BY s.register_no
                """
                params = (filter_date, filter_dept) if filter_dept else (filter_date,)
            cur.execute(query, params)
            records = cur.fetchall()
    finally:
        db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1a73e8")
    hdr_align = Alignment(horizontal='center')

    headers = (['Name','Register No','Department','Days Present']
               if report_type == 'monthly' else
               ['Name','Register No','Department','Date','Time','Status'])
    ws.append(headers)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

    for r in records:
        if report_type == 'monthly':
            ws.append([r['name'], r['register_no'], r['department'], r['days_present']])
        else:
            ws.append([r['name'], r['register_no'], r['department'],
                       r.get('att_date',''), r.get('att_time',''), r.get('status','Absent')])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 3, 35)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'attendance_{report_type}_{filter_date}.xlsx')


# ─── ANALYTICS ROUTE ──────────────────────────────────────────────────────────
@app.route('/admin/analytics')
@admin_required
def admin_analytics():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(DISTINCT date) as total FROM attendance")
            working_days = cur.fetchone()['total'] or 1

            cur.execute("""
                SELECT s.id, s.name, s.register_no, s.department,
                       COUNT(a.id) as days_present,
                       ROUND(COUNT(a.id) / %s * 100, 1) as percentage
                FROM students s
                LEFT JOIN attendance a ON s.id = a.student_id
                GROUP BY s.id, s.name, s.register_no, s.department
                ORDER BY percentage DESC
            """, (working_days,))
            students = cur.fetchall()

            cur.execute("""
                SELECT s.department,
                       COUNT(DISTINCT s.id) as total_students,
                       ROUND(AVG(sub.pct), 1) as avg_percentage
                FROM students s
                LEFT JOIN (
                    SELECT student_id,
                           COUNT(*) / %s * 100 as pct
                    FROM attendance
                    GROUP BY student_id
                ) sub ON s.id = sub.student_id
                GROUP BY s.department
                ORDER BY avg_percentage DESC
            """, (working_days,))
            dept_stats = cur.fetchall()

            cur.execute("""
                SELECT DATE_FORMAT(date, '%%d-%%m') as day,
                       COUNT(*) as count
                FROM attendance
                WHERE date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY date
                ORDER BY date ASC
            """)
            daily_trend = cur.fetchall()

            cur.execute("SELECT COUNT(*) as count FROM students")
            total_students = cur.fetchone()['count']

            cur.execute("SELECT COUNT(DISTINCT student_id) as count FROM attendance")
            ever_attended = cur.fetchone()['count']
    finally:
        db.close()

    # ─── K-Means Clustering ───────────────────────────────
    cluster_data = []
    cluster_summary = {}

    if len(students) >= 50:
        df = pd.DataFrame(students)
        df['percentage']  = df['percentage'].fillna(0)
        df['days_present'] = df['days_present'].fillna(0)

        features = df[['days_present', 'percentage']].values
        scaler   = StandardScaler()
        scaled   = scaler.fit_transform(features)

        kmeans = KMeans(n_clusters=3, init='k-means++',
                        max_iter=300, n_init=10, random_state=42)
        df['cluster'] = kmeans.fit_predict(scaled)

        cluster_means = df.groupby('cluster')['percentage'].mean().sort_values(ascending=False)
        label_map = {
            cluster_means.index[0]: 'High (≥75%)',
            cluster_means.index[1]: 'Medium (50–74%)',
            cluster_means.index[2]: 'Low (<50%)'
        }
        df['cluster_label'] = df['cluster'].map(label_map)

        color_map = {
            'High (≥75%)':     '#28a745',
            'Medium (50–74%)': '#ffc107',
            'Low (<50%)':      '#dc3545'
        }
        df['cluster_color'] = df['cluster_label'].map(color_map)

        cluster_data = df[['name','register_no','department',
                            'days_present','percentage',
                            'cluster_label','cluster_color']].to_dict('records')
        cluster_summary = df.groupby('cluster_label').size().to_dict()

    present_pct = [s for s in students if (s['percentage'] or 0) >= 75]
    medium_pct  = [s for s in students if 50 <= (s['percentage'] or 0) < 75]
    low_pct     = [s for s in students if (s['percentage'] or 0) < 50]

    return render_template('admin/analytics.html',
                           students=students,
                           dept_stats=dept_stats,
                           daily_trend=daily_trend,
                           cluster_data=cluster_data,
                           cluster_summary=cluster_summary,
                           total_students=total_students,
                           ever_attended=ever_attended,
                           working_days=working_days,
                           high_count=len(present_pct),
                           medium_count=len(medium_pct),
                           low_count=len(low_pct))


@app.route('/admin/delete_student/<int:sid>', methods=['POST'])
@admin_required
def delete_student(sid):
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("DELETE FROM attendance WHERE student_id=%s", (sid,))
            cur.execute("DELETE FROM students WHERE id=%s", (sid,))
            db.commit()
    finally:
        db.close()
    return redirect(url_for('admin_students'))


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    return redirect(url_for('admin_login'))


if __name__ == '__main__':
    from pyngrok import ngrok, conf

    ngrok.set_auth_token("350rWPMkOkRtHGSbdMhOfIentHg_2mnX4RD389sgzt3wDQXR6")

    # ✅ Kill ALL existing tunnels before starting fresh
    try:
        tunnels = ngrok.get_tunnels()
        for tunnel in tunnels:
            ngrok.disconnect(tunnel.public_url)
            print(f"Disconnected existing tunnel: {tunnel.public_url}")
    except Exception:
        pass  # No existing tunnels — safe to ignore

    public_url = ngrok.connect(5000)
    print(f"NgrokTunnel: \"{public_url}\" -> \"http://localhost:5000\"")
    app.run(debug=False, host='0.0.0.0', port=5000)